"""
1.数据读取：封装一个可以读取任意Excel文件的方法，可以指定读取的表单(文件名，表单名)
2.数据写入：
    文件名：filename
    表单名：sheetname
    行：row
    列：column
    写入的值：value
"""
import openpyxl
from openpyxl.styles import PatternFill


class HandleExcel:
    def __init__(self, filename, sheetname):
        """
        @param filename: excel文件路径
        @param sheetname: 所要使用的excel表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """
        读取excel数据
        @return: dict
        """
        wk = openpyxl.load_workbook(self.filename)
        sh = wk[self.sheetname]
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        """
        向excel写入数据
        @param row: 写入的行
        @param column: 写入的列
        @param value: 写入的值
        @return:
        """
        wk = openpyxl.load_workbook(self.filename)
        sh = wk[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        wk.save(self.filename)

    def bg_color(self, row, column, color='FF0000', fill_type='none'):
        """
        设置excel背景颜色
        @param row: 行号
        @param column: 列号
        @param color: 背景颜色，默认为红色FF0000
        @param fill_type: 边框样式，默认不改变边框样式
        @return:
        """
        wk = openpyxl.load_workbook(self.filename)
        sh = wk[self.sheetname]
        fill = PatternFill(start_color=color, fill_type=fill_type)
        cell = []
        for i in range(65, 65 + column):
            res = chr(i) + str(row)
            cell.append(res)
        for j in cell:
            sh[j].fill = fill
        wk.save(self.filename)
